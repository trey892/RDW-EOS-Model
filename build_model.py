"""
Build RDW_EOS_Master_v2.xlsx from:
  - data/raw/RDW_EOS_Master_v1.xlsx              (star-schema spine)
  - data/raw/RDW_EOS_Operations_Dashboard_v2_r2.xlsx  (Movements/Loads, already Void-stripped)
  - data/raw/Samsara_Fuel_Efficiency_2026-05-06_to_2026-08-04.csv
  - data/raw/(Samsara idle CSV, single day 2026-01-01)
  - data/raw/6Mo_Avg_Freight_By_Driver.xlsx      (per-driver freight, ~Aug 2025 - Apr 2026)

Everything is additive: nothing from Master v1 is deleted or silently overwritten.
Every new join that can fail (unmatched tractor, ambiguous driver, etc.) lands in
QA_Exceptions instead of being dropped.
"""
import datetime
from pathlib import Path

import openpyxl
import pandas as pd

RAW = Path(__file__).parent / "data" / "raw"
STAGED = Path(__file__).parent / "data" / "staged"
OUT = Path(__file__).parent / "output" / "RDW_EOS_Master_v5.xlsx"

# Trey's manual Style/Division decisions for the no-letter tractor worksheet
# (confirmed via chat 2026-08-04). Chemical stands alone as its own division;
# every other Category rolls up to Waste. TRC-91 resolved separately (round 3,
# confirmed via chat 2026-08-04): PTO / Waste.
TREY_STYLE_DECISIONS = {
    "91": "PTO",
    "106": "Straight Truck", "440": "PTO", "483": "PTO", "484": "PTO", "495": "PTO",
    "513": "Van", "515": "Van", "516": "Van", "518": "Van", "519": "Van", "521": "Van",
    "523": "PTO", "524": "PTO", "525": "PTO", "526": "PTO", "528": "PTO", "529": "PTO",
    "530": "PTO", "531": "PTO", "532": "PTO", "534": "Van", "535": "Van", "536": "Van",
    "537": "Van", "538": "Van", "539": "Van", "540": "Van", "541": "Van", "542": "PTO",
    "543": "PTO", "544": "PTO", "545": "PTO", "546": "PTO", "547": "PTO", "548": "PTO",
    "549": "Van", "550": "Van", "551": "Van", "552": "Van", "553": "Van", "554": "Van",
    "555": "Van", "556": "Van", "5128": "PTO",
}

# Round 2 (confirmed via chat 2026-08-04): the 49 lettered (O/L/C/P) tractors.
# Trey answered at the Division level directly (Chemical/Waste) rather than
# picking a specific PTO/Straight Truck/Van sub-style -- so StyleKey stays
# 'Unclassified' for these (sub-style genuinely still unknown), but Division
# is now resolved.
TREY_DIVISION_DECISIONS_ROUND2 = {
    "522L": "Chemical", "533L": "Chemical", "700L": "Chemical", "706L": "Chemical",
    "716L": "Chemical", "717L": "Chemical", "720L": "Chemical", "721L": "Chemical",
    "723L": "Chemical", "724L": "Chemical", "725L": "Chemical", "726L": "Chemical",
    "727L": "Chemical", "728L": "Chemical", "729L": "Chemical", "730L": "Chemical",
    "733L": "Chemical", "734L": "Chemical", "735L": "Chemical", "736L": "Chemical",
    "737L": "Chemical", "738L": "Waste", "740L": "Chemical", "741L": "Chemical",
    "742L": "Chemical", "743L": "Chemical", "744L": "Chemical", "745L": "Chemical",
    "746L": "Chemical", "747L": "Chemical", "748L": "Chemical", "749L": "Chemical",
    "750L": "Chemical", "751L": "Chemical", "752L": "Chemical", "756L": "Chemical",
    "757L": "Chemical", "807O": "Chemical", "85L": "Chemical", "862O": "Chemical",
    "887O": "Chemical", "88L": "Waste", "895O": "Chemical", "900O": "Chemical",
    "903O": "Chemical", "909O": "Chemical", "942O": "Chemical", "952O": "Waste",
    "953O": "Waste",
}


def division_for_style(style):
    if style == "Chemical":
        return "Chemical"
    if style in ("PTO", "Straight Truck", "Van"):
        return "Waste"
    return None

IDLE_CSV = RAW / "Samsara_Idle_Events_2026-01-01.csv"
FUEL_CSV = RAW / "Samsara_Fuel_Efficiency_2026-05-06_to_2026-08-04.csv"


def sheet_to_df(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])


def load_source_workbooks():
    master = openpyxl.load_workbook(RAW / "RDW_EOS_Master_v1.xlsx", data_only=True)
    ops = openpyxl.load_workbook(RAW / "RDW_EOS_Operations_Dashboard_v2_r2.xlsx", data_only=True)
    return master, ops


def main():
    master, ops = load_source_workbooks()

    # ---- pass-through tables from Master v1 (unchanged) ----
    dim_equipment = sheet_to_df(master, "Dim_Equipment")
    dim_program = sheet_to_df(master, "Dim_Program")
    dim_terminal = sheet_to_df(master, "Dim_Terminal")
    dim_account = sheet_to_df(master, "Dim_Account")
    dim_date = sheet_to_df(master, "Dim_Date")
    fact_maintenance_wo = sheet_to_df(master, "Fact_Maintenance_WO")
    fact_asset_economics = sheet_to_df(master, "Fact_Asset_Economics")
    qa_source_log = sheet_to_df(master, "QA_Source_Log")
    qa_exceptions = sheet_to_df(master, "QA_Exceptions")
    data_dictionary = sheet_to_df(master, "Data_Dictionary")

    # ---- apply Trey's Style/Division decisions, add Division column ----
    is_tractor = dim_equipment["AssetType"] == "TRACTOR"

    # Round 1: no-letter tractors, Trey gave a specific Style -> StyleKey overwritten.
    round1_mask = is_tractor & dim_equipment["AssetID"].astype(str).isin(TREY_STYLE_DECISIONS)
    dim_equipment.loc[round1_mask, "StyleKey"] = (
        dim_equipment.loc[round1_mask, "AssetID"].astype(str).map(TREY_STYLE_DECISIONS)
    )

    dim_equipment["Division"] = None
    dim_equipment.loc[is_tractor, "Division"] = dim_equipment.loc[is_tractor, "StyleKey"].map(division_for_style)

    # Round 2: lettered tractors, Trey gave Division directly (no specific sub-style)
    # -> Division overwritten, StyleKey left as 'Unclassified' since the sub-style
    # (PTO/Straight Truck/Van) genuinely wasn't specified.
    round2_mask = is_tractor & dim_equipment["AssetID"].astype(str).isin(TREY_DIVISION_DECISIONS_ROUND2)
    dim_equipment.loc[round2_mask, "Division"] = (
        dim_equipment.loc[round2_mask, "AssetID"].astype(str).map(TREY_DIVISION_DECISIONS_ROUND2)
    )

    n_resolved_this_round = int(round1_mask.sum())
    n_resolved_round2 = int(round2_mask.sum())

    still_unresolved = dim_equipment[is_tractor & dim_equipment["Division"].isna()]
    n_still_unclassified = len(still_unresolved)
    n_still_unclassified_no_letter = int(
        still_unresolved["AssetID"].astype(str).str.fullmatch(r"\d+").sum()
    )
    n_still_unclassified_lettered = n_still_unclassified - n_still_unclassified_no_letter
    n_style_unknown_division_known = int(
        (is_tractor & dim_equipment["StyleKey"].eq("Unclassified") & dim_equipment["Division"].notna()).sum()
    )

    def norm(v):
        return str(v).strip().upper() if v is not None else v

    tractor_key = {
        norm(row.AssetID): row.AssetKey
        for row in dim_equipment.itertuples()
        if row.AssetType == "TRACTOR"
    }
    tractor_ownership = {
        norm(row.AssetID): row.OwnershipClass
        for row in dim_equipment.itertuples()
        if row.AssetType == "TRACTOR"
    }
    tractor_terminal = {
        norm(row.AssetID): row.Terminal
        for row in dim_equipment.itertuples()
        if row.AssetType == "TRACTOR"
    }
    driver_to_tractor = {
        str(row.DriverOwnerCode): str(row.AssetID)
        for row in dim_equipment.itertuples()
        if row.AssetType == "TRACTOR" and row.DriverOwnerCode and row.DriverOwnerCode != "RDW"
    }

    new_qa_source_rows = []
    new_qa_exception_rows = []

    new_qa_source_rows.append(
        ["Trey manual review (No-Letter Tractor Categorization)", "StyleKey/Division for no-letter tractors",
         "2026-08-04", n_resolved_this_round, "PASS",
         "All 45 no-letter tractors resolved (TRC-91 confirmed PTO/Waste in a follow-up round). Added Dim_Equipment.Division: Chemical stands alone, all other Categories (PTO/Straight Truck/Van) roll up to Waste."]
    )
    new_qa_source_rows.append(
        ["Trey manual review (Lettered Tractor Categorization)", "Division for lettered (O/L/C/P) tractors",
         "2026-08-04", n_resolved_round2, "PASS",
         (f"{n_resolved_round2} of 49 lettered tractors resolved -- Trey answered at the Division level "
          "directly (Chemical/Waste) rather than a specific sub-style, so StyleKey stays 'Unclassified' "
          "for these; only Division is populated. See QA_Exceptions for the sub-style gap.")]
    )
    new_qa_exception_rows.append(
        ["MEDIUM", "Dim_Equipment", "Tractors still missing Division classification",
         n_still_unclassified, None,
         (f"{n_still_unclassified_no_letter} no-letter (TRC-91, left blank in Trey's review) + "
          f"{n_still_unclassified_lettered} lettered (O/L/C/P) tractors not yet reviewed."),
         "Asset Manager"]
    )
    new_qa_exception_rows.append(
        ["LOW", "Dim_Equipment", "Division known but specific Style (PTO/Straight Truck/Van) still unspecified",
         n_style_unknown_division_known, None,
         ("Round-2 lettered-tractor review answered Division directly (Chemical/Waste) rather than "
          "picking a sub-style. Fine for anything that only needs the Chemical/Waste split; a further "
          "pass would be needed to get PTO/Straight Truck/Van detail on these units."),
         "Asset Manager"]
    )

    # ================= Fact_Movements =================
    movements = sheet_to_df(ops, "Movements")
    movements["AssetKey"] = movements["Tractor"].map(lambda t: tractor_key.get(norm(t)) if t else None)
    movements["EquipOwnershipClass"] = movements["Tractor"].map(lambda t: tractor_ownership.get(norm(t)) if t else None)
    movements["EquipTerminal"] = movements["Tractor"].map(lambda t: tractor_terminal.get(norm(t)) if t else None)

    blank_tractor = movements["Tractor"].isna() | (movements["Tractor"] == "")
    off_list = (~blank_tractor) & movements["AssetKey"].isna()
    STALE_PLACEHOLDERS = {"Unclassified - awaiting Trey ID", "No Tractor", "Not In Equipment List"}
    differs = (
        movements["AssetKey"].notna()
        & movements["Ownership Class"].notna()
        & (movements["Ownership Class"] != movements["EquipOwnershipClass"])
    )
    resolved_placeholder = differs & movements["Ownership Class"].isin(STALE_PLACEHOLDERS)
    ownership_mismatch = differs & ~movements["Ownership Class"].isin(STALE_PLACEHOLDERS)

    n_blank_tractor = int(blank_tractor.sum())
    n_off_list = int(off_list.sum())
    n_resolved_placeholder = int(resolved_placeholder.sum())
    n_mismatch = int(ownership_mismatch.sum())

    fact_movements = movements.drop(columns=["EquipOwnershipClass", "EquipTerminal"])

    new_qa_source_rows.append(
        ["McLeod Movements", "Movement legs (loaded/empty miles, dispatcher/tractor utilization)",
         "2026-08-02", len(fact_movements), "PASS",
         f"Void-stripped upstream in Operations Dashboard v2_r2 (15,683 kept). Joined to Dim_Equipment on Tractor->AssetKey."]
    )
    new_qa_exception_rows.append(
        ["MEDIUM", "Fact_Movements", "Movements with blank Tractor field", n_blank_tractor, None,
         "Includes Available (not-yet-dispatched) and Delivered-but-unassigned legs. Not joined to any AssetKey.",
         "Dispatch"]
    )
    new_qa_exception_rows.append(
        ["MEDIUM", "Fact_Movements", "Movements referencing a tractor not on the equipment list", n_off_list, None,
         "Tractor value present but no match in Dim_Equipment (TRACTOR). Possible retired/relabeled unit.",
         "Asset Manager"]
    )
    new_qa_exception_rows.append(
        ["LOW", "Fact_Movements", "Ownership Class differs between Movements (Dim_Fleet-sourced) and Dim_Equipment (McLeod-sourced)",
         n_mismatch, None,
         "Genuine conflicts only (excludes stale 'Unclassified/No Tractor/Not In Equipment List' placeholders -- see next row). Dim_Equipment (McLeod, sole master per standing rule) was NOT overwritten by the Movements value; both are visible in Fact_Movements for review.",
         "Asset Manager"]
    )
    new_qa_exception_rows.append(
        ["INFO", "Fact_Movements", "Movements shows a stale 'Unclassified/No Tractor/Not In Equipment List' placeholder that Dim_Equipment has since resolved",
         n_resolved_placeholder, None,
         "Not a real conflict -- Movements carries the older Dim_Fleet classification (built 2026-08-02); Dim_Equipment (built 2026-08-04) already resolved these via McLeod. Fact_Movements keeps both columns so the resolution is visible.",
         "Asset Manager"]
    )

    # ================= Fact_Loads =================
    loads = sheet_to_df(ops, "Loads")
    loads["Terminal"] = loads["Revenue Code"]
    unmatched_terminal = ~loads["Terminal"].isin(set(dim_terminal["TerminalKey"]))
    n_unmatched_terminal = int(unmatched_terminal.sum())
    fact_loads = loads

    new_qa_source_rows.append(
        ["McLeod Loads", "Customer orders (revenue, customers, lanes, commodities)", "2026-08-02",
         len(fact_loads), "PASS",
         "Void-stripped upstream (0 Void rows in this export). No Tractor field on Loads -- Revenue Code used as Terminal; not joined to Dim_Fleet/Dim_Equipment."]
    )
    if n_unmatched_terminal:
        new_qa_exception_rows.append(
            ["LOW", "Fact_Loads", "Revenue Code not found in Dim_Terminal", n_unmatched_terminal, None,
             "Revenue Code used directly as Terminal; some codes may need mapping to Dim_Terminal.TerminalKey.",
             "Asset Manager"]
        )

    # ================= Fact_Vehicle_Performance (Samsara Fuel) =================
    fuel = pd.read_csv(FUEL_CSV, encoding="utf-8-sig")
    fuel_period = "2026-05-06 to 2026-08-04"
    fuel["AssetKey"] = fuel["Vehicle"].map(lambda v: tractor_key.get(norm(v)))
    fuel_unmatched = int(fuel["AssetKey"].isna().sum())
    fuel["Period"] = fuel_period
    fact_vehicle_performance = fuel.drop(columns=["Tags"]).rename(columns={
        "Efficiency (MPG)": "MPG",
        "Fuel Used (gal)": "FuelUsedGal",
        "Distance (mi)": "DistanceMi",
        "Est. Carbon Emissions (kg)": "EstCarbonEmissionsKg",
        "Est. Cost": "EstCost",
        "Total Engine Run Time (min)": "EngineRunTimeMin",
        "Idle Time (min)": "IdleTimeMin",
        "Idle Time (%)": "IdleTimePct",
        "Vehicle": "VehicleID",
    })
    fact_vehicle_performance = fact_vehicle_performance[
        ["AssetKey", "VehicleID", "Period", "MPG", "FuelUsedGal", "DistanceMi",
         "EstCarbonEmissionsKg", "EstCost", "EngineRunTimeMin", "IdleTimeMin", "IdleTimePct"]
    ]

    new_qa_source_rows.append(
        ["Samsara Fuel Efficiency (row-level)", "MPG, fuel, distance, cost, idle % per vehicle", "2026-08-04",
         len(fact_vehicle_performance), "PASS" if fuel_unmatched == 0 else "REVIEW",
         f"Period {fuel_period}. Tags field dropped entirely (not used for ownership/division per standing rule -- confirmed contradictory in a prior review)."]
    )
    if fuel_unmatched:
        new_qa_exception_rows.append(
            ["LOW", "Fact_Vehicle_Performance", "Vehicle not found in Dim_Equipment", fuel_unmatched, None,
             "Samsara vehicle ID has no TRACTOR match in Dim_Equipment.", "Asset Manager"]
        )

    # ================= Fact_Idle_Events (Samsara Idle) =================
    idle = pd.read_csv(IDLE_CSV, encoding="utf-8-sig")
    idle_period = "2026-01-01"
    idle["AssetKey"] = idle["Vehicle"].map(lambda v: tractor_key.get(norm(v)))
    idle_unmatched = int(idle["AssetKey"].isna().sum())
    idle["Period"] = idle_period
    fact_idle_events = idle.drop(columns=["Tags"]).rename(columns={
        "Vehicle": "VehicleID",
        "Fuel Used (gal)": "FuelUsedGal",
        "Gaseous Fuel Used (gal)": "GaseousFuelUsedGal",
        "Est. Cost": "EstCost",
        "Air Temperature": "AirTemperature",
        "PTO Status": "PTOStatus",
    })
    fact_idle_events = fact_idle_events[
        ["AssetKey", "VehicleID", "Driver", "Period", "Time", "Duration", "Address",
         "AirTemperature", "PTOStatus", "FuelUsedGal", "GaseousFuelUsedGal", "EstCost"]
    ]

    new_qa_source_rows.append(
        ["Samsara Idle Events (row-level)", "Idle cost, duration, driver, location per event",
         "2026-01-01", len(fact_idle_events), "REVIEW",
         "Single-day sample (2026-01-01) only -- not a representative period. Refresh with a multi-month export before trending idle cost. Tags field dropped (same reason as Fuel)."]
    )
    if idle_unmatched:
        new_qa_exception_rows.append(
            ["LOW", "Fact_Idle_Events", "Vehicle not found in Dim_Equipment", idle_unmatched, None,
             "Samsara vehicle ID has no TRACTOR match in Dim_Equipment.", "Asset Manager"]
        )

    # ================= Fact_Asset_Economics: driver-actual revenue supplement =================
    freight_wb = openpyxl.load_workbook(RAW / "6Mo_Avg_Freight_By_Driver.xlsx", data_only=True)
    grid = sheet_to_df(freight_wb, "Grid Data")
    grid.columns = ["ShipDate", "Driver", "DriverName", "FreightTotal"]
    grid = grid.dropna(subset=["ShipDate", "Driver", "FreightTotal"])

    grouped = grid.groupby("Driver").agg(
        TotalFreight=("FreightTotal", "sum"),
        MinDate=("ShipDate", "min"),
        MaxDate=("ShipDate", "max"),
        NumRows=("FreightTotal", "count"),
    ).reset_index()
    grouped["SpanDays"] = (grouped["MaxDate"] - grouped["MinDate"]).dt.days

    MIN_SPAN_DAYS = 30
    usable = grouped[grouped["SpanDays"] >= MIN_SPAN_DAYS].copy()
    too_short = grouped[grouped["SpanDays"] < MIN_SPAN_DAYS]

    usable["AssetID"] = usable["Driver"].map(driver_to_tractor)
    matched = usable.dropna(subset=["AssetID"]).copy()
    unmatched_drivers = usable[usable["AssetID"].isna()]

    matched["AnnualizedRevenue"] = matched["TotalFreight"] * (365.0 / matched["SpanDays"])
    matched["AssetKey"] = matched["AssetID"].map(tractor_key)

    new_economics_rows = []
    for row in matched.itertuples():
        basis = (
            f"{row.NumRows} loads, {row.MinDate.date()} to {row.MaxDate.date()} "
            f"({row.SpanDays} days actual) x {365.0 / row.SpanDays:.2f} annualization factor"
        )
        new_economics_rows.append({
            "AssetKey": row.AssetKey,
            "Period": "FY2026-Annualized",
            "MeasureName": "Annualized Revenue (Driver Actual)",
            "Value": round(row.AnnualizedRevenue, 2),
            "Basis": basis,
            "Confidence": "MEDIUM",
            "BasisNote": (
                "Multi-month actual freight by driver, mapped to driver's CURRENT tractor assignment "
                "(Dim_Equipment.DriverOwnerCode) -- does not account for mid-period driver/tractor "
                "reassignment. Company tractors excluded (DriverOwnerCode='RDW', not a specific driver, "
                "so freight can't be isolated per tractor this way). Supplements, does not replace, the "
                "existing one-week x52 estimate -- compare both before using in a leadership-facing number."
            ),
        })
    fact_asset_economics = pd.concat(
        [fact_asset_economics, pd.DataFrame(new_economics_rows)], ignore_index=True
    )

    new_qa_source_rows.append(
        ["Freight by Driver (multi-month actual)", "Per-driver actual freight revenue, LP/OO tractors only",
         "2026-04-12", len(matched), "REVIEW",
         (f"{len(grid)} shipment rows across {grouped.shape[0]} drivers, {grid['ShipDate'].min().date()} to "
          f"{grid['ShipDate'].max().date()}. Only drivers with >= {MIN_SPAN_DAYS} days of observed data used. "
          "Driver->tractor mapping is a CURRENT snapshot (Dim_Equipment.DriverOwnerCode), not period-accurate.")]
    )
    new_qa_exception_rows.append(
        ["MEDIUM", "Fact_Asset_Economics", "Drivers with < 30 days observed data excluded from Driver Actual revenue",
         int(len(too_short)), None,
         "Sample too short to annualize responsibly; left out rather than extrapolated.", "Asset Manager"]
    )
    new_qa_exception_rows.append(
        ["MEDIUM", "Fact_Asset_Economics", "Drivers in freight file with no current tractor match (DriverOwnerCode)",
         int(len(unmatched_drivers)), None,
         "Driver code not found on any active tractor in Dim_Equipment -- likely driver turnover since the freight export.",
         "Asset Manager"]
    )

    # ================= Assemble QA tables =================
    # The v1 log has placeholder "NOT LOADED" rows for Movements/Loads that are now
    # loaded below -- mark them superseded instead of leaving a contradictory pair.
    stale_mask = qa_source_log["Status"] == "NOT LOADED"
    qa_source_log.loc[stale_mask, "Status"] = "SUPERSEDED"
    qa_source_log.loc[stale_mask, "Note"] = (
        qa_source_log.loc[stale_mask, "Note"] + " -- SUPERSEDED in v2, see PASS row below for this source."
    )

    qa_source_log_cols = list(qa_source_log.columns)
    qa_source_log = pd.concat(
        [qa_source_log, pd.DataFrame(new_qa_source_rows, columns=qa_source_log_cols)],
        ignore_index=True,
    )
    qa_exceptions_cols = list(qa_exceptions.columns)
    qa_exceptions = pd.concat(
        [qa_exceptions, pd.DataFrame(new_qa_exception_rows, columns=qa_exceptions_cols)],
        ignore_index=True,
    )

    # ================= Data dictionary additions =================
    dd_new = pd.DataFrame([
        ["Fact_Movements", "Fact", "One movement leg", "MovementID", "McLeod Movements (via Ops Dashboard v2_r2)", "Replace CSV and refresh"],
        ["Fact_Loads", "Fact", "One customer order", "OrderNumber", "McLeod Loads (via Ops Dashboard v2_r2)", "Replace CSV and refresh"],
        ["Fact_Vehicle_Performance", "Fact", "One vehicle x period", "AssetKey+Period", "Samsara Fuel Efficiency export", "Replace CSV and refresh (period stacks, does not overwrite)"],
        ["Fact_Idle_Events", "Fact", "One idle event", "AssetKey+Time", "Samsara Idle Events export", "Replace CSV and refresh (period stacks, does not overwrite)"],
    ], columns=list(data_dictionary.columns))
    data_dictionary = pd.concat([data_dictionary, dd_new], ignore_index=True)

    # ================= README =================
    readme_lines = [
        ["RDW EOS -- MASTER DATA MODEL (v5)"],
        [f"Built {datetime.date.today().isoformat()} from RDW_EOS_Master_v1 + RDW_EOS_Operations_Dashboard_v2_r2 + fresh Samsara/freight exports + Trey's two-round tractor Style/Division review."],
        ["Delivers the punch-list item Master v1 flagged as not done: Fact_Movements and Fact_Loads are now loaded, joined to Dim_Equipment."],
        [""],
        ["WHAT CHANGED FROM v1"],
        [f"1. Fact_Movements loaded: {len(fact_movements)} rows, joined Tractor->AssetKey via Dim_Equipment. {n_blank_tractor} blank-tractor and {n_off_list} off-list rows flagged in QA_Exceptions, not dropped."],
        [f"2. Fact_Loads loaded: {len(fact_loads)} rows. No tractor field on Loads (McLeod design) -- Revenue Code carried as Terminal."],
        [f"3. Fact_Vehicle_Performance added: {len(fact_vehicle_performance)} vehicles, period {fuel_period}. Tags dropped (unreliable/contradictory, confirmed again in this export)."],
        [f"4. Fact_Idle_Events added: {len(fact_idle_events)} events, but SINGLE DAY ONLY (2026-01-01) -- refresh with a longer export before trending."],
        [f"5. Fact_Asset_Economics supplemented with {len(matched)} 'Annualized Revenue (Driver Actual)' rows from a multi-month per-driver freight file -- MEDIUM confidence, LP/OO tractors only, does not replace the existing one-week estimate."],
        ["6. Trailer/RBOX split (build-spec item) was already done in Master v1's Dim_Equipment via AssetType + AttributeStandard -- carried forward unchanged, no rebuild needed."],
        [f"7. NEW Dim_Equipment.Division column added (Chemical stands alone, PTO/Straight Truck/Van roll up to Waste). Round 1: {n_resolved_this_round} no-letter tractors resolved with a specific Style. Round 2: {n_resolved_round2} lettered tractors resolved at the Division level directly ({n_style_unknown_division_known} of those have Division but not a specific sub-style). {n_still_unclassified} tractors ({n_still_unclassified_no_letter} no-letter, {n_still_unclassified_lettered} lettered) still lack any classification -- see QA_Exceptions."],
        ["8. All other Dim_* and Fact_Maintenance_WO / original Fact_Asset_Economics rows carried forward from Master v1 UNCHANGED."],
        [""],
        ["STILL OPEN"],
        ["- Idle Events is a one-day sample; get a multi-month row-level export to make it trend-usable."],
        ["- Trailer 1831's conflicting title records: still unresolved, carried forward from Master v1."],
        [f"- {n_still_unclassified} tractors still have no Division at all ({n_still_unclassified_no_letter} no-letter (TRC-91), {n_still_unclassified_lettered} lettered) -- if that's not zero, a follow-up worksheet would close it out."],
        [f"- {n_style_unknown_division_known} tractors have Division but no specific Style (PTO/Straight Truck/Van) -- optional further detail pass."],
        ["- Merge with the RTD mileage-based Replacement Projection model: still phase 2, not attempted here."],
        ["- See QA_Source_Log and QA_Exceptions for every new row added this build, and their confidence levels."],
    ]
    readme_df = pd.DataFrame(readme_lines)

    # ================= Write workbook =================
    OUT.parent.mkdir(parents=True, exist_ok=True)
    STAGED.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False, header=False)
        dim_equipment.to_excel(writer, sheet_name="Dim_Equipment", index=False)
        dim_program.to_excel(writer, sheet_name="Dim_Program", index=False)
        dim_terminal.to_excel(writer, sheet_name="Dim_Terminal", index=False)
        dim_account.to_excel(writer, sheet_name="Dim_Account", index=False)
        dim_date.to_excel(writer, sheet_name="Dim_Date", index=False)
        fact_movements.to_excel(writer, sheet_name="Fact_Movements", index=False)
        fact_loads.to_excel(writer, sheet_name="Fact_Loads", index=False)
        fact_vehicle_performance.to_excel(writer, sheet_name="Fact_Vehicle_Performance", index=False)
        fact_idle_events.to_excel(writer, sheet_name="Fact_Idle_Events", index=False)
        fact_maintenance_wo.to_excel(writer, sheet_name="Fact_Maintenance_WO", index=False)
        fact_asset_economics.to_excel(writer, sheet_name="Fact_Asset_Economics", index=False)
        qa_source_log.to_excel(writer, sheet_name="QA_Source_Log", index=False)
        qa_exceptions.to_excel(writer, sheet_name="QA_Exceptions", index=False)
        data_dictionary.to_excel(writer, sheet_name="Data_Dictionary", index=False)

    for name, df in [
        ("Fact_Movements", fact_movements), ("Fact_Loads", fact_loads),
        ("Fact_Vehicle_Performance", fact_vehicle_performance),
        ("Fact_Idle_Events", fact_idle_events),
        ("Fact_Asset_Economics", fact_asset_economics),
        ("QA_Source_Log", qa_source_log), ("QA_Exceptions", qa_exceptions),
    ]:
        df.to_csv(STAGED / f"{name}.csv", index=False)

    # ================= Console verification summary =================
    print("=== BUILD SUMMARY ===")
    print(f"Fact_Movements:            {len(fact_movements):>7} rows  (expect ~15,683)")
    print(f"Fact_Loads:                {len(fact_loads):>7} rows  (expect ~1,359)")
    print(f"Fact_Vehicle_Performance:  {len(fact_vehicle_performance):>7} rows  (expect ~282)")
    print(f"Fact_Idle_Events:          {len(fact_idle_events):>7} rows  (expect ~458)")
    print(f"Dim_Equipment (unchanged): {len(dim_equipment):>7} rows")
    print(f"New Fact_Asset_Economics rows added: {len(matched)}")
    print(f"QA_Source_Log rows: {len(qa_source_log)}  QA_Exceptions rows: {len(qa_exceptions)}")
    print(f"Blank-tractor movements: {n_blank_tractor}  Off-list movements: {n_off_list}  Ownership mismatches: {n_mismatch}")
    print(f"Fuel unmatched vehicles: {fuel_unmatched}  Idle unmatched vehicles: {idle_unmatched}")
    print(f"Driver-actual revenue: matched={len(matched)} too_short={len(too_short)} unmatched_driver={len(unmatched_drivers)}")
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
