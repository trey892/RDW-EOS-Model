"""
Build RDW_EOS_Master_v6.xlsx from RDW_EOS_Master_v5.xlsx (base, unchanged unless noted)
plus two additions:
  1. Fact_Maintenance_WO enriched with GridStatus from the RTA work-order status
     grid (data/raw/RTA_WorkOrderSearchGrid_unintegrated.xlsx) where the WO number
     matches -- 1,242 of 2,368 tractor+trailer+RBOX work orders. NOT used to build
     a "current asset status" table: only 222/2,000 grid rows carry a real
     timestamp and WO number doesn't reliably indicate recency, so a "live status"
     claim from this data would likely be wrong.
  2. Fact_Asset_Economics gets two new per-tractor measures: "Cost Per Total Mile"
     and "Cost Per Loaded Mile" (Total RDW Direct Cost / Annualized Distance, the
     loaded-mile version derived via the movement-based deadhead %). Same
     validated math as the dashboard's Maintenance view. Tractors with
     Annualized Distance below MIN_ANNUAL_MILES are excluded, not given a
     misleading $900/mile figure -- logged to QA_Exceptions instead.
"""
from pathlib import Path

import openpyxl
import pandas as pd

RAW = Path(__file__).parent / "data" / "raw"
STAGED = Path(__file__).parent / "data" / "staged"
BASE = Path(__file__).parent / "output" / "RDW_EOS_Master_v5.xlsx"
GRID_FILE = RAW / "RTA_WorkOrderSearchGrid_unintegrated.xlsx"
OUT = Path(__file__).parent / "output" / "RDW_EOS_Master_v6.xlsx"

MIN_ANNUAL_MILES = 10000


def sheet_df(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])


def main():
    base = openpyxl.load_workbook(BASE, data_only=True)
    grid_wb = openpyxl.load_workbook(GRID_FILE, data_only=True)

    dim_equipment = sheet_df(base, "Dim_Equipment")
    dim_program = sheet_df(base, "Dim_Program")
    dim_terminal = sheet_df(base, "Dim_Terminal")
    dim_account = sheet_df(base, "Dim_Account")
    dim_date = sheet_df(base, "Dim_Date")
    fact_movements = sheet_df(base, "Fact_Movements")
    fact_loads = sheet_df(base, "Fact_Loads")
    fact_vehicle_performance = sheet_df(base, "Fact_Vehicle_Performance")
    fact_idle_events = sheet_df(base, "Fact_Idle_Events")
    fact_maintenance_wo = sheet_df(base, "Fact_Maintenance_WO")
    fact_asset_economics = sheet_df(base, "Fact_Asset_Economics")
    qa_source_log = sheet_df(base, "QA_Source_Log")
    qa_exceptions = sheet_df(base, "QA_Exceptions")
    data_dictionary = sheet_df(base, "Data_Dictionary")

    new_qa_source_rows = []
    new_qa_exception_rows = []

    # ================= 1. Enrich Fact_Maintenance_WO with grid status =================
    grid = sheet_df(grid_wb, "workOrderSearchGridDataGrid")
    grid["WO_norm"] = grid["WO"].astype(str)
    grid_status = grid.drop_duplicates(subset="WO_norm", keep="last").set_index("WO_norm")[["WO Status", "Status", "Asset Status"]]
    grid_status = grid_status.rename(columns={"WO Status": "GridWOStatus", "Status": "GridStatus", "Asset Status": "GridAssetStatus"})

    fact_maintenance_wo["WO_norm"] = fact_maintenance_wo["WorkOrderKey"].astype(str).str.split("-").str[-1]
    fact_maintenance_wo = fact_maintenance_wo.merge(grid_status, left_on="WO_norm", right_index=True, how="left").drop(columns=["WO_norm"])

    n_matched = fact_maintenance_wo["GridWOStatus"].notna().sum()
    n_total = len(fact_maintenance_wo)
    new_qa_source_rows.append([
        "RTA Work Order Search Grid", "Work-order status detail (Breakdown/Waiting on Parts/etc), asset status at close",
        "2026-08-08", n_matched, "REVIEW",
        (f"{n_matched} of {n_total} work orders matched by WO number ({n_total - n_matched} not in this export -- "
         "likely outside its date window). NOT used to build a current/live asset-status table: only 222 of 2,000 "
         "grid rows carry a real Schedule Time, and WO number doesn't reliably order by recency in this export."),
    ])

    # ================= 2. Cost Per Mile measures =================
    econ_wide = fact_asset_economics.pivot_table(index="AssetKey", columns="MeasureName", values="Value", aggfunc="first")
    mv = fact_movements.dropna(subset=["AssetKey"]).copy()
    deadhead_by_asset = mv.groupby("AssetKey").apply(
        lambda g: pd.Series({
            "loaded": g.loc[g["Loaded"] == "Yes", "Distance"].sum(),
            "empty": g.loc[g["Loaded"] == "No", "Distance"].sum(),
        }),
        include_groups=False,
    )
    deadhead_by_asset["deadheadPct"] = deadhead_by_asset["empty"] / (deadhead_by_asset["loaded"] + deadhead_by_asset["empty"]) * 100

    new_econ_rows = []
    n_excluded_low_mileage = 0
    n_no_deadhead = 0
    for row in dim_equipment[dim_equipment["AssetType"] == "TRACTOR"].itertuples():
        ak = row.AssetKey
        if ak not in econ_wide.index:
            continue
        e = econ_wide.loc[ak]
        total_cost = e.get("Total RDW Direct Cost")
        annualized_distance = e.get("Annualized Distance")
        if pd.isna(total_cost) or pd.isna(annualized_distance) or annualized_distance < MIN_ANNUAL_MILES:
            n_excluded_low_mileage += 1
            continue
        cost_per_total_mile = total_cost / annualized_distance
        new_econ_rows.append({
            "AssetKey": ak, "Period": "FY2026-Annualized", "MeasureName": "Cost Per Total Mile",
            "Value": round(cost_per_total_mile, 3), "Basis": "Total RDW Direct Cost / Annualized Distance",
            "Confidence": "MEDIUM", "BasisNote": "Both inputs are themselves annualized/extrapolated figures -- treat as directional.",
        })
        if ak in deadhead_by_asset.index and pd.notna(deadhead_by_asset.loc[ak, "deadheadPct"]):
            deadhead_pct = deadhead_by_asset.loc[ak, "deadheadPct"]
            loaded_miles_est = annualized_distance * (1 - deadhead_pct / 100)
            if loaded_miles_est > 0:
                cost_per_loaded_mile = total_cost / loaded_miles_est
                new_econ_rows.append({
                    "AssetKey": ak, "Period": "FY2026-Annualized", "MeasureName": "Cost Per Loaded Mile",
                    "Value": round(cost_per_loaded_mile, 3),
                    "Basis": "Total RDW Direct Cost / (Annualized Distance x (1 - Deadhead %))",
                    "Confidence": "MEDIUM",
                    "BasisNote": f"Deadhead % ({deadhead_pct:.1f}%) from McLeod Movements export, applied to annualized distance -- a blended estimate, not a direct loaded-mile measurement.",
                })
            else:
                n_no_deadhead += 1
        else:
            n_no_deadhead += 1

    fact_asset_economics = pd.concat([fact_asset_economics, pd.DataFrame(new_econ_rows)], ignore_index=True)

    new_qa_source_rows.append([
        "Cost Per Mile (derived)", "Cost per total mile and cost per loaded mile, per tractor", "2026-08-08",
        len(new_econ_rows), "REVIEW",
        (f"{len(new_econ_rows)} rows added ({len(new_econ_rows) - n_no_deadhead} loaded-mile pairs). "
         "Same MEDIUM-confidence math validated in the dashboard's Maintenance view."),
    ])
    new_qa_exception_rows.append([
        "MEDIUM", "Fact_Asset_Economics", "Tractors excluded from Cost Per Mile (Annualized Distance too low)",
        n_excluded_low_mileage, None,
        f"Annualized Distance below {MIN_ANNUAL_MILES:,} miles/year produces an unreliable, inflated $/mile ratio -- excluded rather than reported.",
        "Asset Manager",
    ])

    # ================= Assemble QA tables =================
    qa_source_log_cols = list(qa_source_log.columns)
    qa_source_log = pd.concat([qa_source_log, pd.DataFrame(new_qa_source_rows, columns=qa_source_log_cols)], ignore_index=True)
    qa_exceptions_cols = list(qa_exceptions.columns)
    qa_exceptions = pd.concat([qa_exceptions, pd.DataFrame(new_qa_exception_rows, columns=qa_exceptions_cols)], ignore_index=True)

    dd_new = pd.DataFrame([
        ["Fact_Maintenance_WO.GridWOStatus/GridStatus/GridAssetStatus", "Fact enrichment", "One work order (same grain as Fact_Maintenance_WO)",
         "WorkOrderKey", "RTA Work Order Search Grid", "Replace CSV and refresh; re-match by WO number"],
    ], columns=list(data_dictionary.columns))
    data_dictionary = pd.concat([data_dictionary, dd_new], ignore_index=True)

    # ================= README =================
    readme_lines = [
        ["RDW EOS -- MASTER DATA MODEL (v6)"],
        ["Built 2026-08-08 from RDW_EOS_Master_v5.xlsx + RTA Work Order Search Grid."],
        [""],
        ["WHAT CHANGED FROM v5"],
        [f"1. Fact_Maintenance_WO enriched with GridWOStatus/GridStatus/GridAssetStatus for {n_matched} of {n_total} work orders (matched by WO number). Richer status detail (Breakdown/Waiting On Parts/Waiting On Tech/etc) than the existing IsOpen flag -- but NOT a live/current-status table (see QA_Source_Log for why)."],
        [f"2. Fact_Asset_Economics gained 'Cost Per Total Mile' and 'Cost Per Loaded Mile' measures -- {len(new_econ_rows)} rows, MEDIUM confidence, same validated math as the dashboard's Maintenance view. {n_excluded_low_mileage} tractors excluded for unreliable low-mileage denominators (see QA_Exceptions)."],
        ["3. All other tables carried forward from v5 UNCHANGED (Dim_Equipment's Division/StyleKey resolution from the two-round manual review is untouched)."],
        [""],
        ["STILL OPEN"],
        ["- PM due-date tracking: still not available in any source data. Needs a dedicated RTA PM Due / Preventive Maintenance Forecast export."],
        ["- Idle Events is still a one-day sample (2026-01-01 only)."],
        ["- Trailer 1831's conflicting title records: still unresolved."],
        ["- 49 lettered tractors have Division but not a specific Style (PTO/Straight Truck/Van) sub-type -- optional further pass."],
        ["- See QA_Source_Log and QA_Exceptions for full detail and confidence levels on everything added this build."],
    ]
    readme_df = pd.DataFrame(readme_lines)

    # ================= Write workbook =================
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False, header=False)
        dim_equipment.to_excel(writer, sheet_name="Dim_Equipment", index=False)
        dim_program.to_excel(writer, sheet_name="Dim_Program", index=False)
        dim_terminal.to_excel(writer, sheet_name="Dim_Terminal", index=False)
        dim_account.to_excel(writer, sheet_name="Dim_Account", index=False)
        dim_date.to_excel(writer, sheet_name="Dim_Date", index=False)
        fact_movements.to_excel(writer, sheet_name="Fact_Movements", index=False)
        fact_loads.to_excel(writer, sheet_name="Fact_Loads", index=False)
        fact_vehicle_performance.to_excel(writer, sheet_name="Fact_Vehicle_Performance", index=False)
        fact_idle_events.to_excel(writer, sheet_name="Fact_Idle_Events", index=False)
        fact_maintenance_wo.to_excel(writer, sheet_name="Fact_Maintenance_WO", index=False)
        fact_asset_economics.to_excel(writer, sheet_name="Fact_Asset_Economics", index=False)
        qa_source_log.to_excel(writer, sheet_name="QA_Source_Log", index=False)
        qa_exceptions.to_excel(writer, sheet_name="QA_Exceptions", index=False)
        data_dictionary.to_excel(writer, sheet_name="Data_Dictionary", index=False)

    print("=== BUILD SUMMARY (v6) ===")
    print(f"Fact_Maintenance_WO: {n_total} rows, {n_matched} matched to grid status ({n_total - n_matched} unmatched)")
    print(f"New Fact_Asset_Economics rows: {len(new_econ_rows)} (excluded {n_excluded_low_mileage} low-mileage tractors)")
    print(f"QA_Source_Log rows: {len(qa_source_log)}  QA_Exceptions rows: {len(qa_exceptions)}")
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
