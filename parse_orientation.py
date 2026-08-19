"""
Parse driver orientation class rosters from Orientation.xlsm -- one sheet per
class date (sheet name IS the date, e.g. "9.1.26", inconsistently zero-padded).
Row 1 has a sync timestamp; row 2 is the header; data rows are one per
orientation slot (ROOM 1-4 or "HOTEL"), most of which are empty placeholders.

Picks the "current" class as the nearest sheet date to today: the next
upcoming one if today falls before a class, otherwise today's own class if
one is scheduled today. Driver rows are those with both First Name and Last
Name populated -- empty ROOM/HOTEL placeholder rows are skipped.

Not part of the daily automated refresh yet -- pulled from the same
"SharePoint Sync" Drive folder as the RTD workbook; wire into daily_refresh.py
once the file-pull step is added to the scheduled routine, same pattern as
RTD_latest.xlsx.
"""
import json
import re
from datetime import date, datetime
from pathlib import Path

from dashboard_filters import is_excluded_terminal

import openpyxl

SRC = Path(__file__).parent / "data" / "raw" / "Orientation_latest.xlsm"
OUT = Path(__file__).parent / "output" / "orientation_data.json"

SHEET_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{2})$")


def sheet_to_date(name):
    m = SHEET_DATE_RE.match(name.strip())
    if not m:
        return None
    month, day, yy = (int(g) for g in m.groups())
    return date(2000 + yy, month, day)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=False)

    dated_sheets = []
    for name in wb.sheetnames:
        d = sheet_to_date(name)
        if d:
            dated_sheets.append((d, name))
    dated_sheets.sort()

    if not dated_sheets:
        raise RuntimeError(f"No date-named sheets found in {SRC}")

    today = date.today()
    upcoming = [(d, n) for d, n in dated_sheets if d >= today]
    class_date, sheet_name = upcoming[0] if upcoming else dated_sheets[-1]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]  # row 2 (0-indexed 1)
    idx = {h: i for i, h in enumerate(header) if h}

    def cell(r, col):
        i = idx.get(col)
        return r[i] if i is not None and i < len(r) else None

    drivers = []
    for r in rows[2:]:
        first = (cell(r, "First Name") or "").strip() if cell(r, "First Name") else ""
        last = (cell(r, "Last Name") or "").strip() if cell(r, "Last Name") else ""
        if not first and not last:
            continue
        ownership_raw = (cell(r, "CO/OO") or "").strip().upper()
        terminal = str(cell(r, "TERMINAL") or "").strip()
        if is_excluded_terminal(terminal):
            continue
        drivers.append({
            "room": str(cell(r, "ROOM") or "").strip(),
            "terminal": terminal,
            "firstName": first,
            "lastName": last,
            "returningDriver": str(cell(r, "Returning Driver") or "").strip(),
            "ownership": ownership_raw,  # "CO" or "OO" (raw, not expanded -- source values inconsistent case)
            "ownerName": str(cell(r, "Owner Name") or "").strip(),
        })

    from collections import Counter
    fleet_counts = Counter(d["ownership"] or "UNKNOWN" for d in drivers)
    terminal_counts = Counter(d["terminal"] or "UNKNOWN" for d in drivers)
    fleets = [
        {"fleet": k, "count": v, "company": fleet_counts.get("CO", 0), "ownerOperator": fleet_counts.get("OO", 0)}
        for k, v in terminal_counts.items()
    ]

    out = {
        "classDate": class_date.isoformat(),
        "classDateLabel": f"{class_date.strftime('%B')} {class_date.day}, {class_date.year}",
        "classWeekday": class_date.strftime("%A"),
        "sheetName": sheet_name,
        "source": "Orientation.xlsm (SharePoint Sync)",
        "checkedAt": datetime.now().isoformat(timespec="seconds"),
        "driverCount": len(drivers),
        "drivers": drivers,
        "fleets": fleets,
    }

    print(f"=== Orientation: {sheet_name} ({class_date.isoformat()}, {class_date.strftime('%A')}) ===")
    print(f"  {len(drivers)} inbound driver(s)")
    for k, v in terminal_counts.most_common():
        print(f"  {k}: {v}")
    print(f"\nWrote -> {OUT}")

    OUT.write_text(json.dumps(out, default=str, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
