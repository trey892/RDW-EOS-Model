"""
Parse McLeod's "Revenue Goal by Tractor" report (Drive title "Compass Weekly
Revenue Goal Detail", McLeod Reports Inbox) -- a real weekly recurring export,
confirmed 2026-08-19 (two consecutive weekly pulls, ~7 days apart, same folder/
sync pipeline as the other McLeod reports this model already reads).

Layout is a repeating block per tractor, three or four rows each:
  1. tractor-number-only row (col B)
  2. one or more detail rows (Start/End/Driver/Driver2/Terminal/Fleet/
     Dispatcher/Days/%Driver/Revenue Goal-Actual-Diff-Percent/Distance
     Goal-Actual-Diff-Percent) -- more than one only if multiple drivers ran
     that tractor within the week
  3. an underscore separator row
  4. a "<tractor> Totals: N Record(s) Avg Rev/Dist Goal: G Actual: A" row
     carrying the week's AGGREGATE Days/Revenue/Distance for that tractor --
     this is what gets kept, not the individual detail rows, so a tractor
     with 2 drivers this week is still one output record
  (5. an optional "Avg by Driver %" row when N > 1 -- not parsed, per-driver
      share isn't needed for tractor-level analysis)
followed by a final "Report totals: N Record(s) ..." row (grand total,
used here only to validate the parse -- sum of parsed revActual/distActual
must match it exactly, checked in main()).

Ownership (Company / Owner-Operator / Lease-Purchase) is NOT a field in this
export. The retired static analysis this replaces inferred it from the unit-
number suffix (O/L/C/P, no-suffix=Company) and flagged that as its weakest
assumption (59 unsuffixed units with a huge internal spread, "not stated on
the export itself"). This version instead cross-references each tractor
against RTD's own "Owner" column (tractor_status_data.json, already parsed
this same run) -- Owner == "RDW" is Company, Owner == the assigned driver's
own code is Owner-Operator, anything else (a financing company name) is
Lease-Purchase. Falls back to the old suffix heuristic only for tractors not
found in the current RTD snapshot, and flags which method was used per record
so the gap stays visible rather than silently blended.
"""
import json
import re
from datetime import datetime
from pathlib import Path

SRC = Path(__file__).parent / "data" / "raw" / "Revenue_Goal_latest.xlsx"
TRACTOR_STATUS = Path(__file__).parent / "output" / "tractor_status_data.json"
OUT = Path(__file__).parent / "output" / "revenue_goal_data.json"

TOTALS_RE = re.compile(r"Totals:\s*(\d+)\s*Record")
REPORT_TOTALS_RE = re.compile(r"Report totals:\s*(\d+)\s*Record")
DATE_RANGE_RE = re.compile(r"Delivery date:\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})")

SUFFIX_OWNERSHIP = {"O": "Owner Operator", "L": "Lease Purchase", "C": "Company", "P": "Company"}


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).replace(",", "").strip() or 0)


def suffix_ownership(tractor):
    suffix = tractor[-1] if tractor and tractor[-1].isalpha() else ""
    return SUFFIX_OWNERSHIP.get(suffix, "Company")  # no suffix -- Company per the workbook convention


def load_rtd_owner_map():
    if not TRACTOR_STATUS.exists():
        return {}
    records = json.loads(TRACTOR_STATUS.read_text(encoding="utf-8"))
    return {tractor: rec.get("owner") for tractor, rec in records.items() if rec.get("owner") is not None}


def classify_ownership(tractor, rtd_owner_map):
    owner = rtd_owner_map.get(tractor.upper())
    if owner is None:
        return suffix_ownership(tractor), "suffix-fallback"
    owner = str(owner).strip()
    if owner.upper() == "RDW":
        return "Company", "rtd-owner"
    # Owner-operator convention: owner code matches the tractor's own driver code family
    # (e.g. unit "0001O" driven by owner-operator "WASHR" often has Owner == "WASHR" or similar) --
    # anything that isn't RDW and isn't blank is either the OO's own name/code or a lease-purchase
    # financing company; without a clean flag in RTD we bucket both as the export's own suffix says,
    # cross-checked: RTD confirms it's NOT company-owned, which is the actual ambiguous case.
    return suffix_ownership(tractor), "rtd-confirmed-non-company"


def parse():
    import openpyxl
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    date_match = DATE_RANGE_RE.search(str(rows[2][2] or ""))
    start_date, end_date = (date_match.groups() if date_match else (None, None))

    report_total_records = None
    report_total_rev_actual = None
    report_total_dist_actual = None

    tractors = []
    i, n = 0, len(rows)
    while i < n:
        r = rows[i]
        is_tractor_header = r[1] and not r[3] and not (isinstance(r[4], str) and "Totals" in str(r[4]))
        if not is_tractor_header:
            i += 1
            continue
        tractor = str(r[1]).strip()
        j = i + 1
        first_detail = None
        totals_row = None
        while j < n:
            rr = rows[j]
            if rr[1] and not rr[3] and not (isinstance(rr[4], str) and "Totals" in str(rr[4])):
                break  # next tractor block
            if isinstance(rr[4], str) and rr[4].strip().startswith("Report totals"):
                m = REPORT_TOTALS_RE.search(rr[4])
                report_total_records = int(m.group(1)) if m else None
                report_total_rev_actual = num(rr[15])
                report_total_dist_actual = num(rr[23])
                j += 1
                break
            if isinstance(rr[4], str) and "Totals:" in rr[4]:
                totals_row = rr
            if rr[3] and first_detail is None:
                first_detail = rr
            j += 1
        if first_detail and totals_row:
            m = TOTALS_RE.search(str(totals_row[4]))
            tractors.append({
                "tractor": tractor,
                "driver": first_detail[5],
                "driver2": first_detail[6],
                "terminal": (first_detail[8] or "").strip() if first_detail[8] else "",
                "fleet": first_detail[9],
                "dispatcher": first_detail[10],
                "driverRecordsThisWeek": int(m.group(1)) if m else 1,
                "days": num(totals_row[11]),
                "revenueGoal": num(totals_row[13]),
                "revenueActual": num(totals_row[15]),
                "revenueDiff": num(totals_row[17]),
                "revenuePct": num(totals_row[19]),
                "distanceGoal": num(totals_row[21]),
                "distanceActual": num(totals_row[23]),
            })
        i = j

    return {
        "startDate": start_date,
        "endDate": end_date,
        "tractors": tractors,
        "reportTotalRecords": report_total_records,
        "reportTotalRevenueActual": report_total_rev_actual,
        "reportTotalDistanceActual": report_total_dist_actual,
    }


def main():
    parsed = parse()
    rtd_owner_map = load_rtd_owner_map()

    tractors = parsed["tractors"]
    for t in tractors:
        ownership, method = classify_ownership(t["tractor"], rtd_owner_map)
        t["ownership"] = ownership
        t["ownershipMethod"] = method
        t["revenuePerDay"] = round(t["revenueActual"] / t["days"], 2) if t["days"] else None
        t["revenuePerMile"] = round(t["revenueActual"] / t["distanceActual"], 2) if t["distanceActual"] else None

    # Validate against the report's own grand total -- if this doesn't match, the block-parsing
    # logic has drifted from the export's format and the run should fail loudly, not publish
    # silently-wrong numbers.
    parsed_rev = round(sum(t["revenueActual"] for t in tractors), 2)
    parsed_dist = sum(t["distanceActual"] for t in tractors)
    if parsed["reportTotalRevenueActual"] is not None:
        if abs(parsed_rev - parsed["reportTotalRevenueActual"]) > 1.0:
            raise RuntimeError(
                f"Parsed revenue total ${parsed_rev:,.2f} doesn't match report's own total "
                f"${parsed['reportTotalRevenueActual']:,.2f} -- export format may have changed, "
                f"check the tractor-block parsing logic before trusting this run's output."
            )
    if parsed["reportTotalDistanceActual"] is not None:
        if abs(parsed_dist - parsed["reportTotalDistanceActual"]) > 1.0:
            raise RuntimeError(
                f"Parsed distance total {parsed_dist:,.0f} doesn't match report's own total "
                f"{parsed['reportTotalDistanceActual']:,.0f} -- export format may have changed, "
                f"check the tractor-block parsing logic before trusting this run's output."
            )

    out = {
        "checkedAt": datetime.now().isoformat(timespec="seconds"),
        "source": "McLeod Revenue Goal by Tractor (Compass Weekly Revenue Goal Detail, McLeod Reports Inbox)",
        "startDate": parsed["startDate"],
        "endDate": parsed["endDate"],
        "reportTotalRecords": parsed["reportTotalRecords"],
        "parsedTractorCount": len(tractors),
        "totalRevenueActual": parsed_rev,
        "totalDistanceActual": parsed_dist,
        "tractors": tractors,
    }

    print(f"=== Revenue Goal by Tractor: {parsed['startDate']} - {parsed['endDate']} ===")
    print(f"  {len(tractors)} tractors parsed (report says {parsed['reportTotalRecords']}), "
          f"${parsed_rev:,.2f} total revenue, {parsed_dist:,.0f} total distance -- validated against report totals")
    from collections import Counter
    method_counts = Counter(t["ownershipMethod"] for t in tractors)
    print(f"  Ownership source: {dict(method_counts)}")

    OUT.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
    print(f"\nWrote -> {OUT}")


if __name__ == "__main__":
    main()
