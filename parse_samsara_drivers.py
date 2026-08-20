"""
Parse the two Samsara scheduled driver reports Trey set up 2026-08-20:
  - "Driver Fuel Mileage and Idle Time (Last 7 Days)" -> MPG, fuel gallons,
    distance, engine idle time, one row per driver.
  - "Driver Safety Scores (Last 30 Days)" -> safety score, one row per driver.

Both are keyed by driver FULL NAME ("JERRY C. PARKS, JR"), not by the driver
code or tractor # the rest of this model uses -- there is no clean join key
on the Samsara side. This matches names against driver_roster_data.json
(parse_driver_roster.py's output, itself sourced from McLeod's own driver
listing) via a normalized "FIRST LAST" key (punctuation/case/generational-
suffix stripped, middle initial dropped since Samsara doesn't reliably carry
one). Per Trey's direction (2026-08-20): best-effort match, but never guess --
every row that doesn't resolve to exactly one roster entry is counted as
unmatched and excluded from the joined output rather than force-matched.
"""
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

FUEL_SRC = Path(__file__).parent / "data" / "raw" / "Samsara_Fuel_Efficiency_latest.xlsx"
SAFETY_SRC = Path(__file__).parent / "data" / "raw" / "Samsara_Safety_Scores_latest.xlsx"
ROSTER_JSON = Path(__file__).parent / "output" / "driver_roster_data.json"
OUT = Path(__file__).parent / "output" / "samsara_driver_data.json"

SRC = FUEL_SRC  # what _run_optional in daily_refresh.py checks for -- if the fuel report exists, run

SUFFIX_RE = re.compile(r"\s*,?\s*\b(JR|SR|II|III|IV)\.?\s*$", re.IGNORECASE)


def normalize_name(full_name):
    """Samsara names come as "LAST, FIRST MIDDLE" or "First Middle Last[, Suffix]" --
    both forms actually reduce to the same FIRST+LAST key logic once suffix/punctuation
    are stripped, since the roster-side key is also just FIRST+LAST (see parse_driver_roster.py)."""
    name = SUFFIX_RE.sub("", str(full_name or "")).strip()
    if "," in name:
        last, _, rest = name.partition(",")
        parts = rest.strip().split()
        first = parts[0] if parts else ""
        key = f"{first} {last.strip()}"
    else:
        parts = name.split()
        if len(parts) >= 2:
            key = f"{parts[0]} {parts[-1]}"
        else:
            key = name
    key = re.sub(r"[^\w\s]", "", key).upper()
    key = re.sub(r"\s+", " ", key).strip()
    return key


def parse_hms_to_minutes(v):
    if not v:
        return None
    s = str(v).strip()
    m = re.match(r"^(\d+):(\d{2}):(\d{2})$", s)
    if not m:
        return None
    h, mi, sec = (int(x) for x in m.groups())
    return round(h * 60 + mi + sec / 60, 1)


def num(v):
    """Samsara leaves some cells as a literal empty string (not blank/None) for drivers with
    no data in the report window -- confirmed 2026-08-20, e.g. a driver who didn't drive in the
    last 7/30 days. Treat that the same as missing, not as a real 0 -- an empty-string safety
    score sorted numerically would otherwise look like the WORST possible score, which is exactly
    backwards from "no data"."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_rows(path):
    # NOT read_only -- these Samsara exports return only column A under read_only mode
    # for reasons unclear (openpyxl quirk with this file's internal structure), confirmed 2026-08-20.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    return header, rows[1:]


def main():
    roster = json.loads(ROSTER_JSON.read_text(encoding="utf-8")) if ROSTER_JSON.exists() else None
    if not roster:
        raise RuntimeError(f"{ROSTER_JSON} not found -- run parse_driver_roster.py first")
    name_index = roster["nameIndex"]
    drivers_by_code = roster["drivers"]

    fuel_header, fuel_rows = read_rows(FUEL_SRC)
    fi = {h: i for i, h in enumerate(fuel_header)}

    records = {}
    fuel_matched, fuel_unmatched = 0, []
    for r in fuel_rows:
        name = r[fi.get("Name", 0)]
        if not name:
            continue
        key = normalize_name(name)
        code = name_index.get(key)
        if not code:
            fuel_unmatched.append(str(name))
            continue
        fuel_matched += 1
        mpg_col = next((h for h in fuel_header if h.startswith("Efficiency")), None)
        fuel_col = next((h for h in fuel_header if h.startswith("Fuel Consumed")), None)
        dist_col = next((h for h in fuel_header if h.startswith("Distance")), None)
        idle_col = next((h for h in fuel_header if h.startswith("Engine Idle")), None)
        records.setdefault(code, {"code": code, "samsaraName": str(name)})
        records[code]["mpg"] = num(r[fi[mpg_col]]) if mpg_col else None
        records[code]["fuelGallons"] = num(r[fi[fuel_col]]) if fuel_col else None
        records[code]["distanceMiles"] = num(r[fi[dist_col]]) if dist_col else None
        records[code]["idleTimeMinutes"] = parse_hms_to_minutes(r[fi[idle_col]]) if idle_col else None

    safety_matched, safety_unmatched = 0, []
    if SAFETY_SRC.exists():
        safety_header, safety_rows = read_rows(SAFETY_SRC)
        si = {h: i for i, h in enumerate(safety_header)}
        score_col = next((h for h in safety_header if "Safety Score" in h), None)
        for r in safety_rows:
            name = r[si.get("Name", 0)]
            if not name:
                continue
            key = normalize_name(name)
            code = name_index.get(key)
            if not code:
                safety_unmatched.append(str(name))
                continue
            score = num(r[si[score_col]]) if score_col else None
            if score is None:
                continue  # "matched a name" but no actual score in this window -- don't count as matched data
            safety_matched += 1
            records.setdefault(code, {"code": code, "samsaraName": str(name)})
            records[code]["safetyScore"] = score

    for code, rec in records.items():
        roster_rec = drivers_by_code.get(code, {})
        rec["fleet"] = roster_rec.get("fleet")
        rec["driverManager"] = roster_rec.get("driverManager")
        rec["rosterName"] = f"{roster_rec.get('firstName', '')} {roster_rec.get('lastName', '')}".strip()

    out = {
        "checkedAt": datetime.now().isoformat(timespec="seconds"),
        "source": "Samsara scheduled reports (Driver Fuel Mileage and Idle Time, Driver Safety Scores), "
                   "joined to McLeod's driver roster by normalized name -- not a McLeod-native join, see module docstring",
        "matchedDriverCount": len(records),
        "fuelReport": {"matched": fuel_matched, "unmatchedCount": len(fuel_unmatched), "unmatchedSample": sorted(set(fuel_unmatched))[:25]},
        "safetyReport": {"matched": safety_matched, "unmatchedCount": len(safety_unmatched), "unmatchedSample": sorted(set(safety_unmatched))[:25]},
        "drivers": list(records.values()),
    }

    print(f"=== Samsara Driver Data ===")
    print(f"  Fuel/Idle: {fuel_matched} matched, {len(fuel_unmatched)} unmatched (of {len(fuel_rows)} rows)")
    print(f"  Safety: {safety_matched} matched, {len(safety_unmatched)} unmatched")
    print(f"  {len(records)} distinct drivers with at least one metric joined")

    OUT.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
    print(f"Wrote -> {OUT}")


if __name__ == "__main__":
    main()
